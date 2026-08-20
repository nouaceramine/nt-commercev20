"""
Customer Debts Routes - Extracted from server.py
Debt tracking, payments (oldest-first), summary, Excel export
"""
from fastapi import APIRouter, HTTPException, Depends
from pydantic import BaseModel
from typing import Optional
from datetime import datetime, timezone
import uuid


from services.balances import customer_debt_aggregates, adjust_customer_mirror, adjust_supplier_mirror, allocate_customer_payment, allocate_supplier_payment


def create_customer_debts_routes(db, get_current_user, get_tenant_admin, require_tenant, CURRENCY="دج") -> dict:
    router = APIRouter(tags=["customer-debts"])

    class CustomerDebtPayment(BaseModel):
        amount: float
        payment_method: str = "cash"
        notes: str = ""

    class SupplierDebtPayment(BaseModel):
        supplier_id: str
        amount: float
        payment_method: str = "cash"

    @router.get("/customers/{customer_id}/debt")
    async def get_customer_debt(customer_id: str, user: dict = Depends(require_tenant)):
        customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
        if not customer:
            raise HTTPException(status_code=404, detail="Customer not found")
        sales = await db.sales.find({"customer_id": customer_id, "debt_amount": {"$gt": 0}}, {"_id": 0}).sort("created_at", -1).to_list(100)
        payments = await db.debt_payments.find({"customer_id": customer_id}, {"_id": 0}).sort("created_at", -1).to_list(100)
        total_debt = sum(s.get("debt_amount", 0) for s in sales)
        return {"customer_id": customer_id, "customer_name": customer.get("name", ""), "total_debt": total_debt, "unpaid_sales": sales, "payment_history": payments}

    @router.post("/customers/{customer_id}/debt/pay")
    async def pay_customer_debt(customer_id: str, payment: CustomerDebtPayment, user: dict = Depends(require_tenant)):
        customer = await db.customers.find_one({"id": customer_id})
        if not customer:
            raise HTTPException(status_code=404, detail="Customer not found")
        # p195: atomic settlement — allocation + payment record + box movement + outbox event commit/abort together
        from config.database import client as _client, main_db as _main_db
        from services.outbox import outbox_write
        async with await _client.start_session() as _tx:
            async with _tx.start_transaction():
                # p64: allocate across open sales via `remaining` (legacy `debt_amount` synced too)
                actual_payment, sales_updated = await allocate_customer_payment(db, customer_id, payment.amount, method=payment.payment_method, session=_tx)
                if actual_payment <= 0:
                    raise HTTPException(status_code=400, detail="Customer has no debt")
                remaining_payment = payment.amount - actual_payment
                now = datetime.now(timezone.utc).isoformat()
                payment_record = {"id": str(uuid.uuid4()), "customer_id": customer_id, "customer_name": customer.get("name", ""), "amount": actual_payment, "payment_method": payment.payment_method, "notes": payment.notes, "sales_updated": sales_updated, "created_at": now, "created_by": user.get("name", "")}
                await db.debt_payments.insert_one(payment_record, session=_tx)
                await adjust_customer_mirror(db, customer_id, total_debt=-actual_payment, balance=-actual_payment, session=_tx)
                # p64: money received must enter a cash box (personal money stays outside boxes)
                if True:  # p68: personal box is a real ledger
                    await db.cash_boxes.update_one({"id": payment.payment_method}, {"$inc": {"balance": actual_payment}, "$set": {"updated_at": now}}, session=_tx)
                    await db.transactions.insert_one({"id": str(uuid.uuid4()), "cash_box_id": payment.payment_method, "type": "income", "amount": actual_payment, "description": f"سداد دين زبون - {customer.get('name', '')}", "reference_type": "debt_payment", "reference_id": payment_record["id"], "created_at": now, "created_by": user.get("name", "")}, session=_tx)
                # p195: outbox → auto journal entry (Dr box / Cr 411)
                await outbox_write(
                    _main_db, "customer.payment_received",
                    {
                        "payment_id": payment_record["id"],
                        "customer_id": customer_id,
                        "customer_name": customer.get("name", ""),
                        "amount": actual_payment,
                        "payment_method": payment.payment_method,
                        "sales_updated": len(sales_updated),
                    },
                    tenant_id=user.get("tenant_id") or "platform",
                    source="customer_debts_routes",
                    session=_tx,
                )
        return {"success": True, "payment_applied": actual_payment, "remaining_from_payment": remaining_payment, "sales_updated": sales_updated}

    @router.post("/supplier-debts/pay")
    async def pay_supplier_debt(payment: SupplierDebtPayment, user: dict = Depends(require_tenant)):
        # p195: atomic settlement — allocation + mirror + box movement + outbox event commit/abort together
        from config.database import client as _client, main_db as _main_db
        from services.outbox import outbox_write
        settlement_id = str(uuid.uuid4())
        async with await _client.start_session() as _tx:
            async with _tx.start_transaction():
                # p64: shared FIFO allocator (same semantics as before)
                amount_applied, updated_purchases = await allocate_supplier_payment(db, payment.supplier_id, payment.amount, method=payment.payment_method, session=_tx)
                if amount_applied <= 0:
                    raise HTTPException(status_code=400, detail="No outstanding debt for this supplier")
                supplier = await db.suppliers.find_one({"id": payment.supplier_id}, session=_tx)
                if supplier:
                    # p62 fix: paying debt reduces what we OWE (balance) — lifetime purchases must not shrink
                    await adjust_supplier_mirror(db, payment.supplier_id, balance=-payment.amount, session=_tx)
                now = datetime.now(timezone.utc).isoformat()
                # p62: personal money lives outside business cash boxes — no box movement
                if True:  # p68: personal box is a real ledger
                    # p64: standard transaction shape (cash_box_id) so /cash ledger shows it
                    await db.transactions.insert_one({"id": str(uuid.uuid4()), "cash_box_id": payment.payment_method, "type": "expense", "amount": payment.amount, "description": f"سداد دين مورد - {supplier['name'] if supplier else payment.supplier_id}", "reference_type": "debt_payment", "reference_id": payment.supplier_id, "settlement_id": settlement_id, "created_at": now, "created_by": user.get("name", "")}, session=_tx)
                    await db.cash_boxes.update_one({"id": payment.payment_method}, {"$inc": {"balance": -payment.amount}, "$set": {"updated_at": now}}, session=_tx)
                # p195: outbox → auto journal entry (Dr 401 / Cr box)
                await outbox_write(
                    _main_db, "supplier.payment_made",
                    {
                        "payment_id": settlement_id,
                        "supplier_id": payment.supplier_id,
                        "supplier_name": supplier.get("name", "") if supplier else "",
                        "amount": payment.amount,
                        "amount_applied": amount_applied,
                        "payment_method": payment.payment_method,
                        "purchases_updated": len(updated_purchases),
                    },
                    tenant_id=user.get("tenant_id") or "platform",
                    source="customer_debts_routes",
                    session=_tx,
                )
        return {"message": "Payment recorded successfully", "amount_paid": payment.amount, "updated_purchases": updated_purchases}

    @router.get("/debts/summary")
    async def get_debts_summary(user: dict = Depends(require_tenant)):
        debts_by_customer = await customer_debt_aggregates(db)
        result = []
        for debt in debts_by_customer:
            customer = await db.customers.find_one({"id": debt["_id"]}, {"_id": 0, "name": 1, "phone": 1})
            if customer:
                result.append({"customer_id": debt["_id"], "customer_name": customer.get("name", "Unknown"), "customer_phone": customer.get("phone", ""), "total_debt": debt["total_debt"], "sales_count": debt["sales_count"]})
        total_outstanding = sum(d["total_debt"] for d in result)
        return {"total_outstanding": total_outstanding, "customers_with_debt": len(result), "debts": sorted(result, key=lambda x: x["total_debt"], reverse=True)}

    @router.get("/debts/export")
    async def export_debts_to_excel(user: dict = Depends(require_tenant)):
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from fastapi.responses import StreamingResponse
        debts_by_customer = await customer_debt_aggregates(db)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Customer Debts"
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        headers = ["#", "اسم الزبون", "رقم الهاتف", "عدد الفواتير", f"إجمالي الدين ({CURRENCY})"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        row_num = 2
        total_debt = 0
        for idx, debt in enumerate(debts_by_customer, 1):
            customer = await db.customers.find_one({"id": debt["_id"]}, {"_id": 0})
            if not customer:
                continue
            ws.cell(row=row_num, column=1, value=idx).border = border
            ws.cell(row=row_num, column=2, value=customer.get("name", "")).border = border
            ws.cell(row=row_num, column=3, value=customer.get("phone", "")).border = border
            ws.cell(row=row_num, column=4, value=debt["sales_count"]).border = border
            cell = ws.cell(row=row_num, column=5, value=debt["total_debt"])
            cell.border = border
            cell.number_format = '#,##0.00'
            total_debt += debt["total_debt"]
            row_num += 1
        ws.cell(row=row_num, column=4, value="الإجمالي:").font = Font(bold=True)
        total_cell = ws.cell(row=row_num, column=5, value=total_debt)
        total_cell.font = Font(bold=True, color="FF0000")
        total_cell.number_format = '#,##0.00'
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 18
        ws.sheet_view.rightToLeft = True
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=debts_{datetime.now().strftime('%Y%m%d')}.xlsx"})

    return router
