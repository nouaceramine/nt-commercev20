"""E-Commerce Hub: Excel/CSV bulk order upload (p243).

EcoManager-style: merchants who receive orders by phone/Excel paste them into
the unified inbox in one shot instead of typing one by one.

  POST /api/ecom/import/orders        — multipart xlsx/csv -> batch import
  GET  /api/ecom/import/template      — ready xlsx template with headers
  GET  /api/ecom/import-batches       — batch history
  GET  /api/ecom/import-batches/{id}  — batch detail (errors included)

Column aliases (first row = header, Arabic or English):
  name:    name / الاسم / اسم الزبون
  phone:   phone / الهاتف / رقم الهاتف / تليفون
  wilaya:  wilaya / الولاية
  city:    city / commune / البلدية / المدينة
  address: address / العنوان
  product: product / المنتج / السلعة
  qty:     qty / quantity / الكمية
  price:   price / السعر
  notes:   notes / ملاحظات

Rows with missing name/phone or bad numbers are SKIPPED and reported per-row;
they never fail the batch. Imported orders go through the same pipeline as
manual entry: duplicate flag (p240, also inside the file itself), COD risk,
network reputation, POS sale mirror, notifications. Tags: excel-import + batch id.
"""
from datetime import datetime, timezone
from typing import Optional
import csv
import io
import logging
import re
import uuid

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse

from config.database import db
from utils.auth import require_tenant
from routes.ecom.constants import require_ecom_feature
from services.application.ecom_order_service import normalize_phone

logger = logging.getLogger(__name__)
router = APIRouter(tags=["E-Commerce Bulk Import"])

MAX_ROWS = 1000

COLUMN_ALIASES = {
    "name":    ["name", "الاسم", "اسم الزبون", "اسم"],
    "phone":   ["phone", "الهاتف", "رقم الهاتف", "تليفون", "الهاتف 1"],
    "wilaya":  ["wilaya", "الولاية"],
    "city":    ["city", "commune", "البلدية", "المدينة"],
    "address": ["address", "العنوان"],
    "product": ["product", "المنتج", "السلعة"],
    "qty":     ["qty", "quantity", "الكمية"],
    "price":   ["price", "السعر"],
    "notes":   ["notes", "ملاحظات", "ملاحظة"],
}


def _map_columns(header: list) -> dict:
    """header cell -> canonical field, case/space-insensitive."""
    mapping = {}
    for i, cell in enumerate(header):
        c = re.sub(r"\s+", " ", str(cell or "").strip().lower())
        for field, aliases in COLUMN_ALIASES.items():
            if c in [a.lower() for a in aliases]:
                mapping[i] = field
                break
    return mapping


def _rows_from_upload(filename: str, raw: bytes):
    """-> (header, rows). Supports .xlsx and .csv (utf-8 / utf-8-sig)."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        text = raw.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        rows = [r for r in reader if any(str(c).strip() for c in r)]
        if not rows:
            raise HTTPException(status_code=400, detail="ملف فارغ")
        return rows[0], rows[1:]
    # default: xlsx
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="صيغة الملف غير مدعومة — ارفعوا xlsx أو csv")
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        vals = ["" if v is None else v for v in row]
        if any(str(v).strip() for v in vals):
            rows.append(vals)
        if len(rows) > MAX_ROWS + 1:
            raise HTTPException(status_code=400, detail=f"الحد الأقصى {MAX_ROWS} صفاً في الدفعة")
    wb.close()
    if not rows:
        raise HTTPException(status_code=400, detail="ملف فارغ")
    return rows[0], rows[1:]


@router.get("/ecom/import/template")
async def bulk_template(user: dict = Depends(require_tenant)):
    await require_ecom_feature(user)
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["الاسم", "الهاتف", "الولاية", "البلدية", "العنوان", "المنتج", "الكمية", "السعر", "ملاحظات"])
    ws.append(["مثال: أحمد بن علي", "0555123456", "16", "الجزائر الوسطى", "حي 200 مسكن", "منتج تجريبي", 1, 1500, ""])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="orders_template.xlsx"'},
    )


@router.post("/ecom/import/orders")
async def bulk_upload(file: UploadFile = File(...), user: dict = Depends(require_tenant)):
    await require_ecom_feature(user)
    raw = await file.read()
    if len(raw) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="الملف أكبر من 5MB")
    header, rows = _rows_from_upload(file.filename or "", raw)
    mapping = _map_columns(header)
    mapped_fields = set(mapping.values())
    if "name" not in mapped_fields or "phone" not in mapped_fields:
        raise HTTPException(status_code=400, detail="عمودا الاسم والهاتف إلزاميان في الترويسة")
    if len(rows) > MAX_ROWS:
        raise HTTPException(status_code=400, detail=f"الحد الأقصى {MAX_ROWS} صفاً في الدفعة")

    batch_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    created, errors = [], []
    seen_in_file = set()

    from services.ecom.duplicate_detector import annotate_order
    from services.cod_risk import calculate_risk_score
    from services.application.ecom_order_service import get_network_trust, reputation_on_create, sync_sale_doc

    def _cell(row, field):
        for i, f in mapping.items():
            if f == field and i < len(row):
                return str(row[i]).strip()
        return ""

    for idx, row in enumerate(rows, start=2):  # row 1 = header
        name = _cell(row, "name")
        phone = normalize_phone(_cell(row, "phone"))
        if not name:
            errors.append({"row": idx, "reason": "الاسم فارغ"})
            continue
        if not phone or len(phone) < 9:
            errors.append({"row": idx, "reason": "هاتف غير صالح"})
            continue
        try:
            qty = max(1, int(float(_cell(row, "qty") or 1)))
            price = max(0.0, float(_cell(row, "price") or 0))
        except ValueError:
            errors.append({"row": idx, "reason": "كمية/سعر غير صالح"})
            continue
        product = _cell(row, "product") or "منتج"
        subtotal = round(qty * price, 2)
        order_id = str(uuid.uuid4())
        ts = datetime.now(timezone.utc).isoformat()
        doc = {
            "id": order_id,
            "order_code": f"XL-{uuid.uuid4().hex[:8].upper()}",
            "channel": "excel",
            "external_id": "",
            "integration_id": None,
            "status": "new",
            "payment_status": "unpaid",
            "payment_method": "cod",
            "customer": {"name": name, "phone": phone, "address": _cell(row, "address"),
                         "city": _cell(row, "city"), "wilaya": _cell(row, "wilaya")},
            "items": [{"name": product, "sku": "", "qty": qty, "price": price, "total": subtotal}],
            "subtotal": subtotal, "shipping_fee": 0, "total": subtotal,
            "notes": _cell(row, "notes"),
            "tags": ["excel-import", f"batch-{batch_id}"],
            "import_batch_id": batch_id,
            "shipping_label_id": None, "tracking_number": None, "courier": None,
            "utm": {}, "utm_source": "",
            "status_history": [{"status": "new", "at": ts, "by": f"excel-import:{batch_id[:8]}"}],
            "created_at": ts, "updated_at": ts,
            "created_by": user.get("id"),
        }
        # in-file duplicate (second+ occurrence of the phone in the same file)
        in_file_dup = phone in seen_in_file
        seen_in_file.add(phone)
        try:
            await annotate_order(db, doc)
        except Exception:  # noqa: BLE001
            pass
        if in_file_dup and not doc.get("duplicate_warning"):
            doc["duplicate_warning"] = True
            doc["duplicate_of"] = {"kind": "file", "code": None, "note": "مكرر داخل الملف نفسه"}
        # COD risk + network reputation (same policy as manual entry)
        try:
            risk = calculate_risk_score(doc, customer_history_count=0, customer_stats={})
            doc["cod_risk"] = risk
            if risk["action"] == "manual_review":
                doc["status"] = "needs_review"
            elif risk["action"] == "confirm_first":
                doc["status"] = "awaiting_confirmation"
        except Exception:  # noqa: BLE001
            pass
        try:
            net = await get_network_trust(phone)
            if net.get("found"):
                doc["network_trust"] = net
        except Exception:  # noqa: BLE001
            pass
        await db.ecom_orders.insert_one(doc)
        try:
            await sync_sale_doc(db, doc)
        except Exception:  # noqa: BLE001
            pass
        try:
            await reputation_on_create(doc, user.get("tenant_id") or "")
        except Exception:  # noqa: BLE001
            pass
        created.append({"row": idx, "order_id": order_id, "order_code": doc["order_code"],
                        "phone": phone, "duplicate_warning": bool(doc.get("duplicate_warning"))})

    batch = {
        "id": batch_id,
        "filename": file.filename or "",
        "total_rows": len(rows),
        "created_count": len(created),
        "skipped_count": len(errors),
        "errors": errors[:200],
        "created_by": user.get("id"),
        "created_at": now,
    }
    await db.ecom_import_batches.insert_one(batch)
    return {"ok": True, "batch_id": batch_id, "created": len(created),
            "skipped": len(errors), "errors": errors[:50], "orders": created[:200]}


@router.get("/ecom/import-batches")
async def list_batches(limit: int = Query(20, ge=1, le=100), user: dict = Depends(require_tenant)):
    await require_ecom_feature(user)
    rows = await db.ecom_import_batches.find({}, {"_id": 0, "errors": 0}).sort("created_at", -1).limit(limit).to_list(limit)
    return {"items": rows}


@router.get("/ecom/import-batches/{batch_id}")
async def get_batch(batch_id: str, user: dict = Depends(require_tenant)):
    await require_ecom_feature(user)
    doc = await db.ecom_import_batches.find_one({"id": batch_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="الدفعة غير موجودة")
    orders = await db.ecom_orders.find(
        {"import_batch_id": batch_id},
        {"_id": 0, "id": 1, "order_code": 1, "status": 1, "customer": 1, "total": 1, "duplicate_warning": 1},
    ).to_list(MAX_ROWS)
    return {"batch": doc, "orders": orders}
